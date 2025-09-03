import re
from pathlib import Path
from datetime import datetime
import sys
import os
import subprocess
import zipfile
import webbrowser
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Настройка UTF-8 вывода
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    messagebox.showerror("Ошибка зависимостей", "Библиотека openpyxl не найдена. Установите ее: pip install openpyxl")
    sys.exit(1)

# ============= НАСТРОЙКИ =============
# Карта статусов и соответствующих им папок
TEMPLATE_STATUSES = {
    "izdato na pregled_GST": "izdato_na_pregled_gst",
    "na uvid_app": "na_uvid_app",
    "za upotrebu_cmm": "za_upotrebu_cmm",
}

# Пути по умолчанию
BASE_DIR = Path(__file__).parent
TEMPLATE_DIR = BASE_DIR / "Template" / "template_tra"
# OUTPUT_DIR = BASE_DIR / "test"
TZ_FILE_PATH = BASE_DIR / "Template" / "TZ.xlsx"

# --- Настройки ячеек и колонок (можно вынести в конфиг) ---
DATE_CELL_ADDR = "C3"
DATE_FMT_TEXT = "%d.%m.%Y"
FOOTER_ANCHOR_NAME = "FooterAnchor"
FIRST_DATA_ROW = 18
COL_RB = 2
COL_BD = 3
COL_NZ = 9
MERGE_BD_FROM, MERGE_BD_TO = 3, 8
MERGE_NZ_FROM, MERGE_NZ_TO = 9, 12
ALLOWED_EXT = { ".pdf", ".docx", ".xlsx", ".xls", ".dwg"}

# --- Регулярные выражения ---
RE_INDEX = re.compile(
    r"\b([IVXLCDM]+)\.(\d+)(?:\.(\d+))?(?:\.(\d+))?([A-Za-zА-Яа-я])?\b",
    re.IGNORECASE
)
DATE_PATTERNS = [
    re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b"),
    re.compile(r"\b\d{4}-\d{2}-\d{2}\b"),
    re.compile(r"\b\d{2}\.\d{2}\.\d{2}\b"),
]

# ---------- БИЗНЕС-ЛОГИКА (ОСНОВНОЙ КОД ОБРАБОТКИ) ----------

def process_files(target_dir: Path, template_path: Path, status_callback, create_archive_flag: bool, delete_files_flag: bool):
    """Основная функция для обработки файлов и создания отчета."""
    try:
        status_callback(f"Загрузка шаблона: {template_path.name}")
        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {template_path}")
        if not target_dir.exists():
            raise FileNotFoundError(f"Папка с файлами не найдена: {target_dir}")
        if not TZ_FILE_PATH.exists():
            status_callback(f"[ПРЕДУПРЕЖДЕНИЕ] Не найден {TZ_FILE_PATH} — 'Назив документа' будет пустым.")

        wb = load_workbook(template_path)
        ws = wb.active

        status_callback("Запись даты...")
        write_date(ws)

        footer_row = get_footer_row_by_name(wb, ws.title, FOOTER_ANCHOR_NAME) or 20
        status_callback(f"Найдена строка футера: {footer_row}")

        status_callback(f"Поиск документов в {target_dir}...")
        files = list_docs(target_dir)
        if not files:
            messagebox.showwarning("Нет файлов", f"В папке {target_dir} не найдено файлов для обработки.")
            return

        status_callback(f"Найдено {len(files)} файлов. Чтение карты индексов...")
        tz_map = build_tz_map_from_xlsx(TZ_FILE_PATH)

        num_files = len(files)
        available_data_rows = footer_row - FIRST_DATA_ROW
        rows_to_insert = 0
        if num_files > available_data_rows:
            rows_to_insert = num_files - available_data_rows

        if rows_to_insert > 0:
            status_callback(f"Вставка {rows_to_insert} строк...")
            insert_rows_and_preserve_footer_merges(ws, footer_row, rows_to_insert)

        new_footer_row = footer_row + rows_to_insert
        status_callback("Заполнение строк данными...")
        final_footer_row = fill_rows(ws, files, tz_map, FIRST_DATA_ROW, new_footer_row)
        
        status_callback("Обновление якоря футера и области печати...")
        update_footer_anchor(wb, ws.title, FOOTER_ANCHOR_NAME, final_footer_row)
        
        last_row = ws.max_row
        ws.print_area = f'B3:P{last_row}'

        wb.template = False
        
        prefix = template_path.stem.replace("-Template", "-")
        saved_path = save_with_increment(wb, target_dir, prefix=prefix)
        
        if create_archive_flag:
            status_callback("Создание ZIP-архива...")
            archive_name = saved_path.with_suffix('').name + "_att.zip"
            archive_path = saved_path.parent / archive_name
            
            try:
                with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_to_add in files:
                        zipf.write(file_to_add, arcname=file_to_add.name)
                
                if delete_files_flag:
                    status_callback("Удаление исходных файлов...")
                    try:
                        for file_to_delete in files:
                            os.remove(file_to_delete)
                        status_callback("Исходные файлы удалены. Открываю папку...")
                    except Exception as e:
                        messagebox.showerror("Ошибка удаления", f"Не удалось удалить исходные файлы: {e}")
                        status_callback("Ошибка удаления файлов.")
                else:
                    status_callback(f"Архив создан. Открываю папку...")

            except Exception as e:
                status_callback(f"Ошибка создания архива: {e}")
                messagebox.showerror("Ошибка архивации", f"Не удалось создать ZIP-архив: {e}")
        else:
             status_callback(f"Готово! Файл сохранен. Открываю папку...")

        try:
            if sys.platform == "win32":
                os.startfile(saved_path.parent)
            elif sys.platform == "darwin":
                subprocess.run(['open', str(saved_path.parent)])
            else:
                subprocess.run(['xdg-open', str(saved_path.parent)])
        except Exception as e:
            messagebox.showwarning("Ошибка", f"Не удалось автоматически открыть папку: {e}")

    except Exception as e:
        status_callback(f"Ошибка: {e}")
        messagebox.showerror("Ошибка выполнения", f"Произошла ошибка:\n{e}")

# ---------- Утилиты (без изменений) ----------

def list_docs(doc_dir: Path):
    return [p for p in sorted(doc_dir.rglob('*'))
            if p.is_file() and p.suffix.lower() in ALLOWED_EXT]

def write_date(ws):
    today = datetime.now().strftime(DATE_FMT_TEXT)
    cell = ws[DATE_CELL_ADDR]
    val = cell.value
    if isinstance(val, str):
        new = val
        for pat in DATE_PATTERNS:
            if pat.search(new):
                new = pat.sub(today, new, count=1)
                break
        else:
            new = today
        cell.value = new
    else:
        cell.value = today

def normalize_key(key: str) -> str:
    key = key.upper()
    replacements = {'A': 'А', 'B': 'Б', 'V': 'В', 'G': 'Г'}
    for lat, cyr in replacements.items():
        key = key.replace(lat, cyr)
    return key

def get_footer_row_by_name(wb, ws_name: str, name: str) -> int | None:
    dn = wb.defined_names.get(name)
    if dn is None: return None
    try:
        destinations = list(dn.destinations)
    except Exception:
        destinations = []
    for sname, ref in destinations:
        s_clean = sname.strip("'") if isinstance(sname, str) else sname
        if s_clean == ws_name:
            coord = str(ref).split("!")[-1].replace("$", "")
            m = re.search(r"\d+", coord)
            if m: return int(m.group(0))
    return None

def update_footer_anchor(wb, ws_name: str, name: str, new_row: int, column_letter: str = "B"):
    ref = f"'{ws_name}'!${column_letter}${new_row}"
    try:
        wb.defined_names.delete(name)
    except Exception:
        pass
    dn_obj = DefinedName(name=name, attr_text=ref)
    try:
        wb.defined_names[name] = dn_obj
    except Exception:
        wb.defined_names.append(dn_obj)

def ensure_row_merges(ws, row, footer_row):
    target_cols_min = min(MERGE_BD_FROM, MERGE_NZ_FROM)
    target_cols_max = max(MERGE_BD_TO, MERGE_NZ_TO)
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        if max_row >= footer_row: continue
        if (min_row <= row <= max_row) and not (max_col < target_cols_min or min_col > target_cols_max):
            to_unmerge.append(str(mr))
    for ref in to_unmerge:
        try:
            ws.unmerge_cells(ref)
        except Exception:
            pass
    rng1 = f"{get_column_letter(MERGE_BD_FROM)}{row}:{get_column_letter(MERGE_BD_TO)}{row}"
    rng2 = f"{get_column_letter(MERGE_NZ_FROM)}{row}:{get_column_letter(MERGE_NZ_TO)}{row}"
    ws.merge_cells(rng1)
    ws.merge_cells(rng2)

def build_tz_map_from_xlsx(xlsx_path: Path) -> dict[str, str]:
    tz_map: dict[str, str] = {}
    if not xlsx_path.exists(): return tz_map
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        max_col = min(ws.max_column, 20)
        for r in range(1, ws.max_row + 1):
            idx_val, idx_col = None, None
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    m = RE_INDEX.search(v)
                    if m:
                        roman, num1, num2, num3, suf = m.groups()
                        suf = suf or ""
                        idx_val = f"{roman.upper()}.{num1}"
                        if num2: idx_val += f".{num2}"
                        if num3: idx_val += f".{num3}"
                        idx_val += suf
                        idx_col = c
                        break
            if not idx_val: continue
            naziv = None
            vC = ws.cell(r, 3).value
            if isinstance(vC, str) and vC.strip():
                naziv = vC.strip()
            else:
                for c in range((idx_col or 1) + 1, max_col + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and len(v.strip()) >= 3:
                        naziv = v.strip()
                        break
            if naziv:
                normalized_key = normalize_key(idx_val)
                if normalized_key not in tz_map:
                    tz_map[normalized_key] = naziv
    return tz_map

def extract_index_from_name(filename: str) -> str | None:
    m = RE_INDEX.search(filename)
    if not m: return None
    roman, num1, num2, num3, suf = m.groups()
    suf = suf or ""
    idx = f"{roman.upper()}.{num1}"
    if num2: idx += f".{num2}"
    if num3: idx += f".{num3}"
    idx += suf
    return idx

def insert_rows_and_preserve_footer_merges(ws, insert_at_row: int, num_rows: int):
    if num_rows <= 0: return
    MAX_COL_TO_COPY = 20
    footer_start_row = insert_at_row
    footer_end_row = ws.max_row
    if footer_end_row < footer_start_row:
        ws.insert_rows(insert_at_row, amount=num_rows)
        return
    footer_snapshot = []
    for r_idx in range(footer_start_row, footer_end_row + 1):
        row_dim = ws.row_dimensions[r_idx]
        row_info = {"height": row_dim.height, "cells": []}
        for c_idx in range(1, MAX_COL_TO_COPY + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            row_info["cells"].append((cell.value, cell._style))
        footer_snapshot.append(row_info)
    footer_merges = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row >= footer_start_row]
    for mr in footer_merges:
        ws.unmerge_cells(str(mr))
    ws.insert_rows(insert_at_row, amount=num_rows)
    new_footer_start_row = footer_start_row + num_rows
    for r_offset, row_info in enumerate(footer_snapshot):
        new_row_num = new_footer_start_row + r_offset
        if row_info["height"] is not None:
            ws.row_dimensions[new_row_num].height = row_info["height"]
        for c_offset, (value, style) in enumerate(row_info["cells"]):
            col_num = 1 + c_offset
            new_cell = ws.cell(row=new_row_num, column=col_num)
            new_cell.value = value
            new_cell._style = style
    for mr in footer_merges:
        mr.shift(0, num_rows)
        ws.merge_cells(str(mr))

def fill_rows(ws, files, tz_map: dict, start_row: int, final_footer_row: int):
    min_col_style, max_col_style = 2, 16
    template_styles = [ws.cell(row=start_row, column=j)._style for j in range(min_col_style, max_col_style + 1)]
    template_row_height = ws.row_dimensions[start_row].height
    const_vals = {
        13: ws.cell(row=start_row, column=13).value,
        14: ws.cell(row=start_row, column=14).value,
        15: ws.cell(row=start_row, column=15).value,
    }
    for i, p in enumerate(files, 1):
        r = start_row + i - 1
        if r >= final_footer_row: continue
        if r > start_row:
            if template_row_height is not None:
                ws.row_dimensions[r].height = template_row_height
            for j_idx, style in enumerate(template_styles):
                ws.cell(row=r, column=min_col_style + j_idx)._style = style
        ensure_row_merges(ws, r, final_footer_row)
        ws.cell(r, COL_RB).value = i
        c = ws.cell(r, COL_BD)
        c.value = p.name
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        idx = extract_index_from_name(p.name)
        base_naziv = tz_map.get(normalize_key(idx), "") if idx else ""
        final_naziv = ""
        if base_naziv:
            prefix = ""
            if "-C-" in p.name.upper(): prefix += "Корективно одржавање. "
            if "_CMM" in p.name.upper(): prefix += "Листа коментара уз документ. "
            final_naziv = prefix + base_naziv
        naziv_cell = ws.cell(r, COL_NZ)
        naziv_cell.value = final_naziv
        naziv_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for col_num, value in const_vals.items():
            ws.cell(row=r, column=col_num).value = value
    return final_footer_row

def save_with_increment(wb, out_dir: Path, prefix="CT-GST-TRA-PRM-"):
    out_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%y%m%d")
    n = 1
    while True:
        out = out_dir / f"{prefix}{today}_{n:02d}.xlsx"
        if not out.exists():
            wb.save(out)
            return out
        n += 1

# ---------- ГРАФИЧЕСКИЙ ИНТЕРФЕЙС (GUI) ----------

def create_transmittal_gui():
    """Создает и управляет GUI для выбора папки и шаблона."""
    root = tk.Tk()
    root.title("Формирование трансмиттала")
    root.geometry("550x640")
    root.resizable(False, False)

    # --- Стилизация ---
    BG_COLOR = "#F4F6F5"
    FRAME_COLOR = "#FFFFFF"
    BUTTON_COLOR = "#4CAF50"
    BUTTON_ACTIVE_COLOR = "#45a049"
    TEXT_COLOR = "#333333"
    DISABLED_TEXT_COLOR = "#aaaaaa"
    STATUS_BAR_COLOR = "#E0E0E0"
    FONT_NORMAL = ("Segoe UI", 10)
    FONT_BOLD = ("Segoe UI", 11, "bold")
    FONT_LABEL = ("Segoe UI", 9)
    FONT_HELP_TEXT = ("Segoe UI", 8)

    root.config(bg=BG_COLOR)

    style = ttk.Style(root)
    style.theme_use('clam')

    style.configure("TButton", background=BUTTON_COLOR, foreground="white", font=FONT_BOLD, bordercolor=BUTTON_COLOR, lightcolor=BUTTON_COLOR, darkcolor=BUTTON_COLOR, padding=(10, 8))
    style.map("TButton", background=[('active', BUTTON_ACTIVE_COLOR)], foreground=[('active', 'white')])
    style.configure("TMenubutton", background="white", foreground=TEXT_COLOR, font=FONT_NORMAL, arrowcolor=TEXT_COLOR, bordercolor=STATUS_BAR_COLOR)
    style.configure("TFrame", background=BG_COLOR)
    style.configure("TLabel", background=BG_COLOR, foreground=TEXT_COLOR, font=FONT_NORMAL)
    style.configure("Header.TLabel", font=FONT_BOLD, background=FRAME_COLOR)
    style.configure("Status.TLabel", background=STATUS_BAR_COLOR, foreground=TEXT_COLOR, padding=5, font=("Segoe UI", 9))
    style.configure("Card.TFrame", background=FRAME_COLOR)
    style.configure("TCheckbutton", background=FRAME_COLOR, font=FONT_NORMAL, foreground=TEXT_COLOR)
    style.map("TCheckbutton", foreground=[('disabled', DISABLED_TEXT_COLOR)])
    style.configure("TRadiobutton", background=FRAME_COLOR, font=FONT_NORMAL, foreground=TEXT_COLOR)
    style.map("TRadiobutton", background=[('active', BG_COLOR)])


    # --- Переменные ---
    selected_folder = tk.StringVar()
    selected_status_key = tk.StringVar(value=list(TEMPLATE_STATUSES.keys())[0])
    selected_template_key = tk.StringVar()
    should_create_archive = tk.BooleanVar(value=True)
    should_delete_files = tk.BooleanVar(value=False)
    
    templates_map = {}

    # --- Функции-обработчики ---
    def open_github(event=None):
        webbrowser.open_new("https://github.com/Dun4ev/toir_plan_checklist")

    def update_template_options(*args):
        nonlocal templates_map
        status_dir_name = TEMPLATE_STATUSES.get(selected_status_key.get())
        if not status_dir_name:
            return

        templates_path = TEMPLATE_DIR / status_dir_name
        templates_map.clear()

        if templates_path.is_dir():
            for f in templates_path.glob("*.xltx"):
                key_name = f.stem.replace("-Template", "").replace("CT-", "").replace("-PRM", "")
                if "XXX" in key_name:
                    key_name = "Общий (XXX)"
                else:
                    key_name = f"{key_name.split('-')[1]} ({key_name.split('-')[0]})"
                templates_map[key_name] = f.name
        
        menu = template_menu["menu"]
        menu.delete(0, "end")
        
        if not templates_map:
            template_menu.config(state=tk.DISABLED)
            selected_template_key.set("")
            return
        
        template_menu.config(state=tk.NORMAL)
        for key in templates_map.keys():
            menu.add_command(label=key, command=tk._setit(selected_template_key, key))
        
        folder_path = selected_folder.get()
        default_key = "Общий (XXX)"
        if folder_path:
            folder_name_upper = Path(folder_path).name.upper()
            available_abbrs = []
            for template_filename in templates_map.values():
                parts = template_filename.split('-')
                if len(parts) > 1 and parts[1].upper() != "XXX":
                    available_abbrs.append(parts[1].upper())
            
            found_template = False
            for abbr in sorted(available_abbrs, key=len, reverse=True):
                if f"_{abbr}" in folder_name_upper or f"-{abbr}" in folder_name_upper:
                    for key, filename in templates_map.items():
                        if f"-{abbr}-" in filename.upper():
                            selected_template_key.set(key)
                            found_template = True
                            break
                if found_template:
                    break
            
            if not found_template:
                if default_key in templates_map:
                    selected_template_key.set(default_key)
        else:
            if default_key in templates_map:
                selected_template_key.set(default_key)
            else:
                selected_template_key.set(list(templates_map.keys())[0] if templates_map else "")

    def toggle_delete_option():
        if should_create_archive.get():
            delete_check.config(state=tk.NORMAL)
        else:
            delete_check.config(state=tk.DISABLED)
            should_delete_files.set(False)

    def select_folder():
        folder_path = filedialog.askdirectory(title="Выберите папку с документами")
        if folder_path:
            selected_folder.set(folder_path)
            folder_display_label.config(text=f"...{folder_path[-50:]}")
            update_template_options()

    def run_processing():
        target_dir = selected_folder.get()
        if not target_dir:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите папку с документами.")
            return
        
        status_dir_name = TEMPLATE_STATUSES.get(selected_status_key.get())
        template_file_name = templates_map.get(selected_template_key.get())

        if not status_dir_name or not template_file_name:
            messagebox.showerror("Ошибка", "Не удалось определить путь к шаблону. Проверьте выбор статуса и шаблона.")
            return

        template_path = TEMPLATE_DIR / status_dir_name / template_file_name

        run_button.config(state=tk.DISABLED)
        def status_update(message):
            status_label.config(text=message)
            root.update_idletasks()

        process_files(Path(target_dir), template_path, status_update, should_create_archive.get(), should_delete_files.get())
        run_button.config(state=tk.NORMAL)

    # --- Компоновка ---
    main_frame = ttk.Frame(root, padding=(15, 10))
    main_frame.pack(fill=tk.BOTH, expand=True)

    # Блок 1: Выбор папки
    folder_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    folder_card.pack(fill=tk.X, pady=5)
    ttk.Label(folder_card, text="1. Выберите папку с документами", style="Header.TLabel").pack(anchor="w")
    folder_display_label = ttk.Label(folder_card, text="(не выбрана)", font=FONT_LABEL, foreground="#757575", background=FRAME_COLOR)
    folder_display_label.pack(anchor="w", pady=(5, 10))
    ttk.Button(folder_card, text="Выбрать папку...", command=select_folder, style="TButton").pack(anchor="w")

    # Блок 2: Выбор статуса отправки
    status_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    status_card.pack(fill=tk.X, pady=5)
    ttk.Label(status_card, text="2. Выберите статус отправки", style="Header.TLabel").pack(anchor="w", pady=(0, 5))
    
    for status_text in TEMPLATE_STATUSES.keys():
        rb = ttk.Radiobutton(status_card, text=status_text, variable=selected_status_key, value=status_text, style="TRadiobutton")
        rb.pack(anchor="w", padx=5)

    # Блок 3: Выбор шаблона
    template_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    template_card.pack(fill=tk.X, pady=5)
    ttk.Label(template_card, text="3. Выберите компанию (шаблон)", style="Header.TLabel").pack(anchor="w")
    
    info_text = ("Подсказка: шаблон выбирается автоматически, если имя папки содержит (GST, TER и т.д.).")
    info_label = ttk.Label(template_card, text=info_text, font=FONT_HELP_TEXT, foreground="#757575", background=FRAME_COLOR, justify=tk.LEFT)
    info_label.pack(anchor='w', pady=(5, 10))

    template_menu = ttk.OptionMenu(template_card, selected_template_key, "", style="TMenubutton")
    template_menu.pack(fill=tk.X)
    template_menu.config(state=tk.DISABLED)

    # Блок 4: Запуск
    run_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    run_card.pack(fill=tk.X, pady=5)
    
    archive_check = ttk.Checkbutton(run_card, text="Создать ZIP-архив с вложениями", variable=should_create_archive, style="TCheckbutton", command=toggle_delete_option)
    archive_check.pack(anchor="w")

    delete_check = ttk.Checkbutton(run_card, text="Удалить исходные файлы после архивации", variable=should_delete_files, style="TCheckbutton")
    delete_check.pack(anchor="w", padx=(20, 0), pady=(0, 15))

    run_button = ttk.Button(run_card, text="Сформировать отчет", command=run_processing, style="TButton")
    run_button.pack(ipady=10, fill=tk.X)

    # --- Нижняя панель (статус-бар и ссылка) ---
    bottom_frame = tk.Frame(root, bg=STATUS_BAR_COLOR)
    bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)

    status_label = ttk.Label(bottom_frame, text="Ожидание...", style="Status.TLabel", anchor="w")
    status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

    link_label = tk.Label(bottom_frame, text="🔗 GitHub", fg="blue", cursor="hand2", bg=STATUS_BAR_COLOR, font=("Segoe UI", 8, "underline"))
    link_label.pack(side=tk.RIGHT, padx=10)
    link_label.bind("<Button-1>", open_github)

    # --- Инициализация и привязки ---
    selected_status_key.trace_add("write", update_template_options)
    toggle_delete_option()
    update_template_options()

    root.mainloop()

if __name__ == "__main__":
    create_transmittal_gui()