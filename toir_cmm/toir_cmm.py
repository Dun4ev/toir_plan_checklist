import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
import tkinter as tk
from tkinter import filedialog

# === НАСТРОЙКИ ===
TEMPLATE_PATH = Path("Template/CommentSheet_Template.xltx")
DATE_FMT = "dd.mm.yyyy"
# Путь к файлу с данными для поиска
TZ_FILE_PATH = Path("Template/TZ.xlsx")

# Регулярка на извлечение полезных фрагментов из имени (пример)
RE_SECTION = re.compile(r"(GMS\d+)", re.IGNORECASE)
RE_I_CODE  = re.compile(r"\b((?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zА-Яа-я])?)\b", re.IGNORECASE)

def find_description_in_tz_file(lookup_key: str) -> str | None:
    """
    Ищет ключ в колонке B файла TZ.xlsx и возвращает
    соответствующее значение из колонки C.
    Поддерживает транслитерацию для кириллических букв в конце ключа.
    """
    if not TZ_FILE_PATH.exists():
        print(f"  - [ERROR] Файл с данными не найден: {TZ_FILE_PATH}")
        return None

    # Карта для транслитерации
    translit_map = {'a': 'а', 'b': 'б', 'v': 'в', 'g': 'г'}
    
    # Создаем список ключей для поиска
    keys_to_find = {lookup_key.strip()}
    
    # Проверяем, заканчивается ли ключ на одну из латинских букв
    if lookup_key and lookup_key[-1] in translit_map:
        # Создаем альтернативный ключ с кириллической буквой
        cyrillic_key = lookup_key[:-1] + translit_map[lookup_key[-1]]
        keys_to_find.add(cyrillic_key)
        print(f"  - [INFO] Обнаружен возможный транслит. Ищем ключи: {keys_to_find}")

    try:
        wb = load_workbook(TZ_FILE_PATH, data_only=True)
        ws = wb['gen_cl'] 
        
        # Ищем точное совпадение в колонке B по любому из ключей
        for row in ws.iter_rows(values_only=True):
            cell_value = str(row[1]).strip() if row[1] else ""
            if cell_value in keys_to_find:
                return row[2] # Возвращаем описание из колонки C
        
        return None # Если ничего не найдено
    except Exception as e:
        print(f"  - [ERROR] Ошибка при чтении файла {TZ_FILE_PATH}: {e}")
        return None

def ensure_named_range(ws, wb, cell, name):
    """Создаёт именованный диапазон, если ещё не существует."""
    existing = set(wb.defined_names.keys())
    if name not in existing:
        dn = DefinedName(name=name, attr_text=f"'{ws.title}'!{cell.coordinate}")
        wb.defined_names.append(dn)

def fill_basic_fields(wb, report_name: str):
    """Заполнить D1/D4 через именованные диапазоны (если их нет — пишем прямо)."""
    ws = wb.active
    dn_map = dict(wb.defined_names.items())

    # ReportName -> D1
    if "ReportName" in dn_map:
        dests = dn_map["ReportName"].destinations
        for sheet, coord in dests:
            ws_target = wb[sheet] if isinstance(sheet, str) else sheet
            ws_target[coord].value = report_name
    else:
        ws["D1"].value = report_name
        ensure_named_range(ws, wb, ws["D1"], "ReportName")

    # CreatedDate -> D4
    today = datetime.now()
    if "CreatedDate" in dn_map:
        for sheet, coord in dn_map["CreatedDate"].destinations:
            ws_target = wb[sheet] if isinstance(sheet, str) else sheet
            cell = ws_target[coord]
            cell.value = today
            cell.number_format = DATE_FMT
    else:
        ws["D4"].value = today
        ws["D4"].number_format = DATE_FMT
        ensure_named_range(ws, wb, ws["D4"], "CreatedDate")

def fill_extra_fields(wb, report_name: str):
    """
    Извлечь код из имени файла, найти его в TZ.xlsx и записать
    соответствующее описание из колонки C в ячейку ExtraField1 (D6).
    """
    ws = wb.active
    icode_match = RE_I_CODE.search(report_name)

    extra_value = "Код не найден в имени файла"
    if icode_match:
        icode = icode_match.group(1)
        print(f"  Найден код в имени файла: {icode}")
        description = find_description_in_tz_file(icode)
        
        if description:
            print(f"  Найдено описание в TZ.xlsx: \"{description}\"")
            extra_value = description
        else:
            print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Код '{icode}' не найден в файле {TZ_FILE_PATH}")
            extra_value = f"ОПИСАНИЕ ДЛЯ {icode} НЕ НАЙДЕНО"
    else:
        print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Код раздела не найден в имени файла: {report_name}")

    # Вставляем найденное значение в ячейку
    dn_map = dict(wb.defined_names.items())
    if "ExtraField1" in dn_map:
        for sheet, coord in dn_map["ExtraField1"].destinations:
            ws_target = wb[sheet] if isinstance(sheet, str) else sheet
            ws_target[coord].value = extra_value
    else:
        # Предполагаем, что плейсхолдер находится в D6, как в оригинальном комментарии
        ws["D6"].value = extra_value

def make_cmm_for_report(report_path: Path):
    """Создает CMM файл для одного отчета."""
    stem = report_path.stem
    cmm_name = f"{stem}_CMM.xlsx"
    cmm_path = report_path.with_name(cmm_name)

    if cmm_path.exists():
        print(f"[ПРОПУСК] Файл уже существует: {cmm_path.name}")
        return

    print(f"Обработка: {report_path.name}")
    try:
        wb = load_workbook(TEMPLATE_PATH)
        wb.template = False
        fill_basic_fields(wb, stem)
        fill_extra_fields(wb, stem)
        wb.save(cmm_path)
        print(f"[OK] Создан файл: {cmm_path.name}")
    except Exception as e:
        print(f"[ОШИБКА] Не удалось обработать {report_path.name}: {e}")

def main():
    """Главная функция для пакетной обработки."""
    # Создаем корневое окно Tkinter, которое нам не нужно показывать
    root = tk.Tk()
    root.withdraw()

    # Открываем диалог выбора папки
    print("Пожалуйста, выберите папку с файлами отчетов (.docx, .pdf)...")
    search_dir = filedialog.askdirectory(
        title="Выберите папку с файлами отчетов (.docx, .pdf)"
    )

    # Если пользователь закрыл диалог, выходим
    if not search_dir:
        print("Папка не выбрана. Завершение работы.")
        return
        
    search_path = Path(search_dir)

    print(f"Запуск пакетной обработки в директории: {search_path.resolve()}")
    
    # Ищем файлы .docx и .pdf
    docx_files = list(search_path.glob("**/*.docx"))
    pdf_files = list(search_path.glob("**/*.pdf"))
    files_to_process = docx_files + pdf_files

    if not files_to_process:
        print(f"В директории '{search_path}' и ее подпапках не найдены файлы .docx или .pdf.")
        return

    processed_files = 0
    for doc_file in files_to_process:
        if doc_file.name.startswith("CT-DR-"):
            make_cmm_for_report(doc_file)
            processed_files += 1
            
    print(f"Обработка завершена. Всего найдено файлов (.docx, .pdf): {len(files_to_process)}. Обработано (с префиксом CT-DR-): {processed_files}.")

if __name__ == "__main__":
    main()
