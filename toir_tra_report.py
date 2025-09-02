import re
from pathlib import Path
from datetime import datetime
import sys
import os
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Настройка UTF-8 вывода
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    messagebox.showerror("Ошибка зависимостей", "Библиотека openpyxl не найдена. Установите ее: pip install openpyxl")
    sys.exit(1)

# ============= НАСТРОЙКИ =============
# Карта выбора шаблонов
TEMPLATES = {
    "Общий (XXX)": "CT-XXX-TRA-PRM-Template.xltx",
    "Gastrans (GST)": "CT-GST-TRA-PRM-Template.xltx",
    "Termoinženjering (TER)": "CT-TER-TRA-PRM-Template.xltx",
}

# Пути по умолчанию
BASE_DIR = Path(__file__).parent
TEMPLATE_DIR = BASE_DIR / "Template" / "template_tra"
# OUTPUT_DIR = BASE_DIR / "test"
TZ_FILE_PATH = BASE_DIR / "Template" / "TZ.xlsx"

# --- Настройки ячеек и колонок (можно вынести в конфиг) ---
DATE_CELL_ADDR = "C3"
DATE_FMT_TEXT = "%d.%m.%Y"
FOOTER_ANCHOR_NAME = "FooterAnchor"
FIRST_DATA_ROW = 18
COL_RB = 2
COL_BD = 3
COL_NZ = 9
MERGE_BD_FROM, MERGE_BD_TO = 3, 8
MERGE_NZ_FROM, MERGE_NZ_TO = 9, 12
ALLOWED_EXT = { ".pdf", ".docx", ".xlsx", ".xls", ".dwg"}

# --- Регулярные выражения ---
RE_INDEX = re.compile(
    r"\b([IVXLCDM]+)\.(\d+)(?:\.(\d+))?(?:\.(\d+))?([A-Za-zА-Яа-я])?\b",
    re.IGNORECASE
)
DATE_PATTERNS = [
    re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b"),
    re.compile(r"\b\d{4}-\d{2}-\d{2}\b"),
    re.compile(r"\b\d{2}\.\d{2}\.\d{2}\b"),
]

# ---------- БИЗНЕС-ЛОГИКА (ОСНОВНОЙ КОД ОБРАБОТКИ) ----------

def process_files(target_dir: Path, template_path: Path, status_callback):
    """Основная функция для обработки файлов и создания отчета."""
    try:
        status_callback(f"Загрузка шаблона: {template_path.name}")
        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {template_path}")
        if not target_dir.exists():
            raise FileNotFoundError(f"Папка с файлами не найдена: {target_dir}")
        if not TZ_FILE_PATH.exists():
            status_callback(f"[ПРЕДУПРЕЖДЕНИЕ] Не найден {TZ_FILE_PATH} — 'Назив документа' будет пустым.")

        wb = load_workbook(template_path)
        ws = wb.active

        status_callback("Запись даты...")
        write_date(ws)

        footer_row = get_footer_row_by_name(wb, ws.title, FOOTER_ANCHOR_NAME) or 20
        status_callback(f"Найдена строка футера: {footer_row}")

        status_callback(f"Поиск документов в {target_dir}...")
        files = list_docs(target_dir)
        if not files:
            messagebox.showwarning("Нет файлов", f"В папке {target_dir} не найдено файлов для обработки.")
            return

        status_callback(f"Найдено {len(files)} файлов. Чтение карты индексов...")
        tz_map = build_tz_map_from_xlsx(TZ_FILE_PATH)

        num_files = len(files)
        available_data_rows = footer_row - FIRST_DATA_ROW
        rows_to_insert = 0
        if num_files > available_data_rows:
            rows_to_insert = num_files - available_data_rows

        if rows_to_insert > 0:
            status_callback(f"Вставка {rows_to_insert} строк...")
            insert_rows_and_preserve_footer_merges(ws, footer_row, rows_to_insert)

        new_footer_row = footer_row + rows_to_insert
        status_callback("Заполнение строк данными...")
        final_footer_row = fill_rows(ws, files, tz_map, FIRST_DATA_ROW, new_footer_row)
        
        status_callback("Обновление якоря футера и области печати...")
        update_footer_anchor(wb, ws.title, FOOTER_ANCHOR_NAME, final_footer_row)
        
        last_row = ws.max_row
        ws.print_area = f'B3:P{last_row}'

        wb.template = False
        
        prefix = template_path.stem.replace("-Template", "-")
        saved_path = save_with_increment(wb, target_dir, prefix=prefix)
        
        status_callback(f"Готово! Файл сохранен: {saved_path}")
        if messagebox.askyesno("Успех", f"Отчет успешно создан:\n{saved_path}\n\nОткрыть папку с файлом?"):
            try:
                if sys.platform == "win32":
                    os.startfile(saved_path.parent)
                elif sys.platform == "darwin":
                    subprocess.run(['open', str(saved_path.parent)])
                else:
                    subprocess.run(['xdg-open', str(saved_path.parent)])
            except Exception as e:
                messagebox.showwarning("Ошибка", f"Не удалось открыть папку: {e}")

    except Exception as e:
        status_callback(f"Ошибка: {e}")
        messagebox.showerror("Ошибка выполнения", f"Произошла ошибка:\n{e}")

# ---------- Утилиты (без изменений) ----------

def list_docs(doc_dir: Path):
    return [p for p in sorted(doc_dir.rglob('*'))
            if p.is_file() and p.suffix.lower() in ALLOWED_EXT]

def write_date(ws):
    today = datetime.now().strftime(DATE_FMT_TEXT)
    cell = ws[DATE_CELL_ADDR]
    val = cell.value
    if isinstance(val, str):
        new = val
        for pat in DATE_PATTERNS:
            if pat.search(new):
                new = pat.sub(today, new, count=1)
                break
        else:
            new = today
        cell.value = new
    else:
        cell.value = today

def normalize_key(key: str) -> str:
    key = key.upper()
    replacements = {'A': 'А', 'B': 'Б', 'V': 'В', 'G': 'Г'}
    for lat, cyr in replacements.items():
        key = key.replace(lat, cyr)
    return key

def get_footer_row_by_name(wb, ws_name: str, name: str) -> int | None:
    dn = wb.defined_names.get(name)
    if dn is None: return None
    try:
        destinations = list(dn.destinations)
    except Exception:
        destinations = []
    for sname, ref in destinations:
        s_clean = sname.strip("'") if isinstance(sname, str) else sname
        if s_clean == ws_name:
            coord = str(ref).split("!")[-1].replace("$", "")
            m = re.search(r"\d+", coord)
            if m: return int(m.group(0))
    return None

def update_footer_anchor(wb, ws_name: str, name: str, new_row: int, column_letter: str = "B"):
    ref = f"'{ws_name}'!${column_letter}${new_row}"
    try:
        wb.defined_names.delete(name)
    except Exception:
        pass
    dn_obj = DefinedName(name=name, attr_text=ref)
    try:
        wb.defined_names[name] = dn_obj
    except Exception:
        wb.defined_names.append(dn_obj)

def ensure_row_merges(ws, row, footer_row):
    target_cols_min = min(MERGE_BD_FROM, MERGE_NZ_FROM)
    target_cols_max = max(MERGE_BD_TO, MERGE_NZ_TO)
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        if max_row >= footer_row: continue
        if (min_row <= row <= max_row) and not (max_col < target_cols_min or min_col > target_cols_max):
            to_unmerge.append(str(mr))
    for ref in to_unmerge:
        try:
            ws.unmerge_cells(ref)
        except Exception:
            pass
    rng1 = f"{get_column_letter(MERGE_BD_FROM)}{row}:{get_column_letter(MERGE_BD_TO)}{row}"
    rng2 = f"{get_column_letter(MERGE_NZ_FROM)}{row}:{get_column_letter(MERGE_NZ_TO)}{row}"
    ws.merge_cells(rng1)
    ws.merge_cells(rng2)

def build_tz_map_from_xlsx(xlsx_path: Path) -> dict[str, str]:
    tz_map: dict[str, str] = {}
    if not xlsx_path.exists(): return tz_map
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        max_col = min(ws.max_column, 20)
        for r in range(1, ws.max_row + 1):
            idx_val, idx_col = None, None
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    m = RE_INDEX.search(v)
                    if m:
                        roman, num1, num2, num3, suf = m.groups()
                        suf = suf or ""
                        idx_val = f"{roman.upper()}.{num1}"
                        if num2: idx_val += f".{num2}"
                        if num3: idx_val += f".{num3}"
                        idx_val += suf
                        idx_col = c
                        break
            if not idx_val: continue
            naziv = None
            vC = ws.cell(r, 3).value
            if isinstance(vC, str) and vC.strip():
                naziv = vC.strip()
            else:
                for c in range((idx_col or 1) + 1, max_col + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and len(v.strip()) >= 3:
                        naziv = v.strip()
                        break
            if naziv:
                normalized_key = normalize_key(idx_val)
                if normalized_key not in tz_map:
                    tz_map[normalized_key] = naziv
    return tz_map

def extract_index_from_name(filename: str) -> str | None:
    m = RE_INDEX.search(filename)
    if not m: return None
    roman, num1, num2, num3, suf = m.groups()
    suf = suf or ""
    idx = f"{roman.upper()}.{num1}"
    if num2: idx += f".{num2}"
    if num3: idx += f".{num3}"
    idx += suf
    return idx

def insert_rows_and_preserve_footer_merges(ws, insert_at_row: int, num_rows: int):
    if num_rows <= 0: return
    MAX_COL_TO_COPY = 20
    footer_start_row = insert_at_row
    footer_end_row = ws.max_row
    if footer_end_row < footer_start_row:
        ws.insert_rows(insert_at_row, amount=num_rows)
        return
    footer_snapshot = []
    for r_idx in range(footer_start_row, footer_end_row + 1):
        row_dim = ws.row_dimensions[r_idx]
        row_info = {"height": row_dim.height, "cells": []}
        for c_idx in range(1, MAX_COL_TO_COPY + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            row_info["cells"].append((cell.value, cell._style))
        footer_snapshot.append(row_info)
    footer_merges = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row >= footer_start_row]
    for mr in footer_merges:
        ws.unmerge_cells(str(mr))
    ws.insert_rows(insert_at_row, amount=num_rows)
    new_footer_start_row = footer_start_row + num_rows
    for r_offset, row_info in enumerate(footer_snapshot):
        new_row_num = new_footer_start_row + r_offset
        if row_info["height"] is not None:
            ws.row_dimensions[new_row_num].height = row_info["height"]
        for c_offset, (value, style) in enumerate(row_info["cells"]):
            col_num = 1 + c_offset
            new_cell = ws.cell(row=new_row_num, column=col_num)
            new_cell.value = value
            new_cell._style = style
    for mr in footer_merges:
        mr.shift(0, num_rows)
        ws.merge_cells(str(mr))

def fill_rows(ws, files, tz_map: dict, start_row: int, final_footer_row: int):
    min_col_style, max_col_style = 2, 16
    template_styles = [ws.cell(row=start_row, column=j)._style for j in range(min_col_style, max_col_style + 1)]
    template_row_height = ws.row_dimensions[start_row].height
    const_vals = {
        13: ws.cell(row=start_row, column=13).value,
        14: ws.cell(row=start_row, column=14).value,
        15: ws.cell(row=start_row, column=15).value,
    }
    for i, p in enumerate(files, 1):
        r = start_row + i - 1
        if r >= final_footer_row: continue
        if r > start_row:
            if template_row_height is not None:
                ws.row_dimensions[r].height = template_row_height
            for j_idx, style in enumerate(template_styles):
                ws.cell(row=r, column=min_col_style + j_idx)._style = style
        ensure_row_merges(ws, r, final_footer_row)
        ws.cell(r, COL_RB).value = i
        c = ws.cell(r, COL_BD)
        c.value = p.name
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        idx = extract_index_from_name(p.name)
        base_naziv = tz_map.get(normalize_key(idx), "") if idx else ""
        final_naziv = ""
        if base_naziv:
            prefix = ""
            if "-C-" in p.name.upper(): prefix += "Корективно одржавање. "
            if "_CMM" in p.name.upper(): prefix += "Листа коментара уз документ. "
            final_naziv = prefix + base_naziv
        naziv_cell = ws.cell(r, COL_NZ)
        naziv_cell.value = final_naziv
        naziv_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for col_num, value in const_vals.items():
            ws.cell(row=r, column=col_num).value = value
    return final_footer_row

def save_with_increment(wb, out_dir: Path, prefix="CT-GST-TRA-PRM-"):
    out_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%y%m%d")
    n = 1
    while True:
        out = out_dir / f"{prefix}{today}_{n:02d}.xlsx"
        if not out.exists():
            wb.save(out)
            return out
        n += 1

# ---------- ГРАФИЧЕСКИЙ ИНТЕРФЕЙС (GUI) ----------

def create_transmittal_gui():
    """Создает и управляет GUI для выбора папки и шаблона."""
    root = tk.Tk()
    root.title("Формирование трансмиттала")
    root.geometry("600x300")

    # Переменные для хранения выбора
    selected_folder = tk.StringVar()
    selected_template_key = tk.StringVar(value=list(TEMPLATES.keys())[0])

    # --- Функции-обработчики ---
    def select_folder():
        folder_path = filedialog.askdirectory(title="Выберите папку с документами")
        if folder_path:
            selected_folder.set(folder_path)
            folder_label.config(text=folder_path)

            # Автоматический выбор шаблона по имени папки
            folder_name_upper = Path(folder_path).name.upper()
            default_key = "Общий (XXX)"

            # Итерация по ключам, кроме общего, для поиска совпадений
            for key in TEMPLATES:
                if key == default_key:
                    continue

                match = re.search(r'\((.*?)\)', key)
                if match:
                    abbreviation = match.group(1).upper()
                    if f"_{abbreviation}" in folder_name_upper or f"-{abbreviation}" in folder_name_upper:
                        selected_template_key.set(key)
                        break
            else:  # Если совпадений не найдено
                selected_template_key.set(default_key)

    def run_processing():
        target_dir = selected_folder.get()
        if not target_dir:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите папку с документами.")
            return

        template_name = TEMPLATES[selected_template_key.get()]
        template_path = TEMPLATE_DIR / template_name

        # Блокируем кнопку, чтобы избежать повторного запуска
        run_button.config(state=tk.DISABLED)
        
        def status_update(message):
            status_label.config(text=message)
            root.update_idletasks()

        process_files(Path(target_dir), template_path, status_update)
        
        # Возвращаем кнопку в активное состояние
        run_button.config(state=tk.NORMAL)

    # --- Виджеты ---
    frame = ttk.Frame(root, padding="10")
    frame.pack(fill=tk.BOTH, expand=True)

    # Выбор папки
    ttk.Label(frame, text="Папка с документами:").pack(pady=5)
    folder_label = ttk.Label(frame, text="(не выбрана)", relief="sunken", padding=5)
    folder_label.pack(fill=tk.X, pady=2)
    ttk.Button(frame, text="Выбрать папку...", command=select_folder).pack(pady=5)

    # Выбор шаблона
    ttk.Label(frame, text="Выберите компанию (шаблон):").pack(pady=10)
    template_menu = ttk.OptionMenu(frame, selected_template_key, selected_template_key.get(), *TEMPLATES.keys())
    template_menu.pack(fill=tk.X, pady=2)

    # Кнопка запуска
    run_button = ttk.Button(frame, text="Сформировать отчет", command=run_processing)
    run_button.pack(pady=20, ipady=10)

    # Статус-бар
    status_label = ttk.Label(root, text="Ожидание...", relief="sunken", anchor="w", padding=5)
    status_label.pack(side=tk.BOTTOM, fill=tk.X)

    root.mainloop()

if __name__ == "__main__":
    create_transmittal_gui()