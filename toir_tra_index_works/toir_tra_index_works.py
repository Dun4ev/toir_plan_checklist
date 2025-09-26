import re
import shutil
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# === НАСТРОЙКИ ===
# Путь к файлу-справочнику, построенный относительно расположения скрипта
script_dir = Path(__file__).resolve().parent
TZ_FILE_PATH = script_dir / "Template" / "TZ_glob.xlsx"
# Имя листа в справочнике
TZ_SHEET_NAME = "gen_cl"
# Колонки в справочнике
TZ_LOOKUP_COL = "B"  # Колонка с индексами (I.7.5)
TZ_SUFFIX_COL = "G"  # Колонка с суффиксами для папки (KBV)
TZ_RESERVED_COL = "H"  # Колонка с кодом Reserved (01, 02 и т.п.)

# Регулярное выражение для извлечения ключа группировки для C-файлов (напр., II.2.6-00-C)
RE_C_GROUPING_KEY = re.compile(
    r"(\b(?:(?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zА-Яа-я])?)(?:-\d{2}-C))\b",
    re.IGNORECASE
)
# Регулярное выражение для извлечения обычного ключа группировки (напр., I.7.5-00-1G)
RE_GROUPING_KEY = re.compile(
    r"(\b(?:(?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zА-Яа-я])?)(?:-\d{2}-\w{1,2}))\b",
    re.IGNORECASE
)
# Регулярное выражение для извлечения индекса для поиска (напр., I.7.5)
RE_INDEX_CODE = re.compile(
    r"(\b(?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zА-Яа-я])?)\b",
    re.IGNORECASE
)

# Карта для транслитерации кириллицы в латиницу для имен папок
CYRILLIC_TO_LATIN = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g',
    'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G'
}

def transliterate_cyrillic_to_latin(text: str) -> str:
    """Конвертирует кириллические символы в латинские по карте."""
    for cyr, lat in CYRILLIC_TO_LATIN.items():
        text = text.replace(cyr, lat)
    return text

def extract_reserved_value(grouping_key: str) -> str | None:
    """Извлекает код Reserved из имени группы (вторая часть после индекса)."""
    parts = grouping_key.split('-')
    if len(parts) < 2:
        return None

    candidate = parts[1].strip()
    if not candidate:
        return None

    if candidate.isdigit():
        return candidate.zfill(2)

    return candidate.upper()

def find_suffix_in_tz_file(lookup_key: str, reserved_code: str | None = None) -> str | None:
    """
    Ищет индекс в колонке B листа 'gen_cl' и возвращает значение из колонки TZ_SUFFIX_COL.
    При наличии кода Reserved дополнительно фильтрует строки по колонке TZ_RESERVED_COL.
    """
    normalized_lookup = lookup_key.strip()
    if not normalized_lookup:
        return None

    if not TZ_FILE_PATH.exists():
        print(f"  - [ОШИБКА] Файл-справочник не найден: {TZ_FILE_PATH}")
        return None

    # Карта для транслитерации латиницы в кириллицу для поиска
    translit_map = {'a': 'а', 'b': 'б', 'v': 'в', 'g': 'г'}

    keys_to_find = {normalized_lookup.lower()}
    last_char = normalized_lookup[-1]

    if last_char in translit_map:
        cyrillic_key = normalized_lookup[:-1] + translit_map[last_char]
        keys_to_find.add(cyrillic_key.lower())

    normalized_reserved = None
    if reserved_code:
        reserved_raw = reserved_code.strip()
        if reserved_raw:
            normalized_reserved = reserved_raw.zfill(2) if reserved_raw.isdigit() else reserved_raw.upper()

    try:
        wb = load_workbook(TZ_FILE_PATH, data_only=True)
        if TZ_SHEET_NAME not in wb.sheetnames:
            print(f"  - [ОШИБКА] Лист '{TZ_SHEET_NAME}' не найден в файле {TZ_FILE_PATH}")
            return None

        ws = wb[TZ_SHEET_NAME]

        lookup_col_idx = ord(TZ_LOOKUP_COL.upper()) - ord('A') + 1
        suffix_col_idx = ord(TZ_SUFFIX_COL.upper()) - ord('A') + 1
        reserved_col_idx = ord(TZ_RESERVED_COL.upper()) - ord('A') + 1

        fallback_suffix: str | None = None
        for row in ws.iter_rows(values_only=True):
            row_lookup = row[lookup_col_idx - 1]
            cell_value = str(row_lookup).strip().lower() if row_lookup else ""
            if cell_value not in keys_to_find:
                continue

            suffix_cell = row[suffix_col_idx - 1] if len(row) >= suffix_col_idx else None
            suffix = str(suffix_cell).strip() if suffix_cell else None
            if not suffix:
                continue

            reserved_cell = row[reserved_col_idx - 1] if len(row) >= reserved_col_idx else None
            normalized_row_reserved = None
            if reserved_cell is not None:
                reserved_str = str(reserved_cell).strip()
                if reserved_str:
                    normalized_row_reserved = reserved_str.zfill(2) if reserved_str.isdigit() else reserved_str.upper()

            if normalized_reserved:
                if normalized_row_reserved == normalized_reserved:
                    return suffix
                if fallback_suffix is None:
                    fallback_suffix = suffix
            else:
                return suffix

        return fallback_suffix
    except Exception as e:
        print(f"  - [ОШИБКА] Ошибка при чтении файла {TZ_FILE_PATH}: {e}")
        return None


def main():
    """Главная функция для сортировки файлов по папкам."""
    # --- Блок выбора папки ---
    # Для быстрого тестирования: раскомментируйте строку ниже и укажите путь
    # source_dir = r"test/toir_tra_index_works"
    source_dir = r"D:\CT DOO\CT_docs - 01_Maintenance\03_Report_base\01_Processing\05_TRA_SUB_app\2025\09.September\1"

    # # Для обычной работы: закомментируйте строку выше и используйте диалог выбора папки
    # if 'source_dir' not in locals():
    #     root = tk.Tk()
    #     root.withdraw()
    #     print("Пожалуйста, выберите папку с файлами для сортировки...")
    #     source_dir = filedialog.askdirectory(title="Выберите папку для сортировки")

    if not source_dir:
        print("Папка не выбрана. Завершение работы.")
        return
        
    source_path = Path(source_dir)
    print(f"Сканирование директории: {source_path.resolve()}")

    # 1. Группируем все файлы по ключу
    files_by_key = defaultdict(list)
    all_files = [p for p in source_path.rglob('*') if p.is_file()]

    for file_path in all_files:
        # Сначала ищем более специфичный ключ для C-файлов
        c_match = RE_C_GROUPING_KEY.search(file_path.name)
        if c_match:
            key = c_match.group(1)
            files_by_key[key].append(file_path)
            continue  # Файл уже сгруппирован, переходим к следующему

        # Если не C-файл, ищем обычный ключ
        match = RE_GROUPING_KEY.search(file_path.name)
        if match:
            key = match.group(1)
            files_by_key[key].append(file_path)

    if not files_by_key:
        print("Не найдено файлов, соответствующих шаблону имен для сортировки.")
        return

    print(f"Найдено {len(files_by_key)} групп файлов для сортировки.")

    # 2. Обрабатываем каждую группу
    for key, file_paths in files_by_key.items():
        print(f"\n--- Обработка группы: {key} ---")

        folder_name: str
        # Проверяем, является ли ключ C-ключом
        if key.upper().endswith("-C"):
            folder_name = transliterate_cyrillic_to_latin(key)
            print(f"  - Группа C-файлов. Имя папки: {folder_name}")
        else:
            # Стандартная логика с поиском суффикса
            index_match = RE_INDEX_CODE.search(key)
            if not index_match:
                print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Не удалось извлечь индекс из ключа '{key}'. Пропуск группы.")
                continue

            index_code = index_match.group(1)
            reserved_code = extract_reserved_value(key)
            print(f"  - Поиск суффикса для индекса: {index_code}")
            if reserved_code:
                print(f"  - Код Reserved из имени: {reserved_code}")

            suffix = find_suffix_in_tz_file(index_code, reserved_code)
            if not suffix:
                if reserved_code:
                    print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Суффикс для индекса '{index_code}' и Reserved '{reserved_code}' не найден в {TZ_FILE_PATH}. Пропуск группы.")
                else:
                    print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Суффикс для индекса '{index_code}' не найден в {TZ_FILE_PATH}. Пропуск группы.")
                continue

            print(f"  - Найден суффикс: '{suffix}'")
            latin_key = transliterate_cyrillic_to_latin(key)
            folder_name = f"{latin_key}_{suffix}"

        # Создаем папку и перемещаем файлы
        dest_dir = source_path / folder_name
        
        print(f"  - Создание папки: {dest_dir.name}")
        dest_dir.mkdir(exist_ok=True)

        # 4. Перемещаем файлы
        for file_path in file_paths:
            dest_file = dest_dir / file_path.name
            try:
                print(f"    - Перемещение: {file_path.name}")
                shutil.move(str(file_path), str(dest_file))
            except Exception as e:
                print(f"    - [ОШИБКА] Не удалось переместить {file_path.name}: {e}")

    print("\nОбработка завершена.")

if __name__ == "__main__":
    main()